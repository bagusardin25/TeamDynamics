"""
Document Analysis Service — extracts text from uploaded documents
and uses LLM to generate structured analysis for simulation setup.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass
from pathlib import Path

from models.document import (
    DocumentAnalysisError,
    DocumentAnalysisResult,
    DocumentTeamSource,
)

logger = logging.getLogger(__name__)


class DocumentExtractionError(ValueError):
    """Raised when an uploaded file cannot be decoded or parsed safely."""


class DocumentRosterEvidenceError(ValueError):
    """Raised when a schema-valid roster cannot be grounded in its document."""


@dataclass(frozen=True)
class DocumentedRosterRow:
    name: str
    role: str
    context: str
    evidence: str


_PDF_ROW_Y_TOLERANCE = 1.5
_PDF_DEFAULT_FONT_SIZE = 8.5
_PDF_AVERAGE_GLYPH_WIDTH_EM = 0.5
_PDF_COLUMN_START_GAP_EM = 6.0
_PDF_COLUMN_END_GAP_EM = 1.5
_PDFTextFragment = (
    tuple[float, float, str]
    | tuple[float, float, str, float]
)


def _pdf_row_text(cells: list[tuple[float, str, float]]) -> str:
    fields: list[str] = []
    current_text = ""
    previous_x: float | None = None
    previous_end: float | None = None
    previous_font_size = _PDF_DEFAULT_FONT_SIZE

    for x, text, font_size in sorted(cells):
        estimated_end = x + (
            max(len(text), 1)
            * font_size
            * _PDF_AVERAGE_GLYPH_WIDTH_EM
        )
        starts_new_field = (
            previous_x is not None
            and previous_end is not None
            and x - previous_x
            > max(previous_font_size, font_size)
            * _PDF_COLUMN_START_GAP_EM
            and x - previous_end
            > max(previous_font_size, font_size)
            * _PDF_COLUMN_END_GAP_EM
        )

        if starts_new_field:
            fields.append(current_text)
            current_text = text
        else:
            current_text = " ".join(
                part for part in (current_text, text) if part
            )

        previous_x = x
        previous_end = max(previous_end or estimated_end, estimated_end)
        previous_font_size = font_size

    if current_text:
        fields.append(current_text)

    return " | ".join(fields)


def _pdf_layout_rows(
    fragments: list[_PDFTextFragment],
) -> list[str]:
    normalized_fragments: list[tuple[float, float, str, float]] = []
    for fragment in fragments:
        x, y, text = fragment[:3]
        if not text or not text.strip():
            continue
        font_size = (
            float(fragment[3])
            if len(fragment) == 4 and fragment[3]
            else _PDF_DEFAULT_FONT_SIZE
        )
        normalized_fragments.append(
            (float(x), float(y), " ".join(text.split()), font_size)
        )

    ordered_fragments = sorted(
        normalized_fragments,
        key=lambda fragment: (-fragment[1], fragment[0]),
    )

    rows: list[str] = []
    current_y: float | None = None
    current_cells: list[tuple[float, str, float]] = []

    for x, y, text, font_size in ordered_fragments:
        if current_y is not None and abs(current_y - y) > _PDF_ROW_Y_TOLERANCE:
            rows.append(_pdf_row_text(current_cells))
            current_cells = []
            current_y = y
        elif current_y is None:
            current_y = y

        current_cells.append((x, text, font_size))

    if current_cells:
        rows.append(_pdf_row_text(current_cells))

    return rows


async def extract_text_from_file(content: bytes, filename: str) -> str:
    """Extract text content from uploaded files (PDF, DOCX, TXT, CSV, XLSX)."""
    ext = Path(filename).suffix.lower()

    if ext == ".txt":
        try:
            return content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise DocumentExtractionError("Could not decode TXT file as UTF-8.") from exc

    elif ext == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise DocumentExtractionError("Could not decode CSV file as UTF-8.") from exc
        try:
            reader = csv.reader(io.StringIO(text))
            rows = []
            for row in reader:
                rows.append(" | ".join(row))
            return "\n".join(rows)
        except csv.Error as exc:
            logger.warning("CSV parsing failed for %s", filename, exc_info=True)
            raise DocumentExtractionError("Could not parse CSV file.") from exc

    elif ext == ".pdf":
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(io.BytesIO(content))
            pages = []
            for page in reader.pages:
                fragments: list[_PDFTextFragment] = []

                def collect_pdf_fragment(
                    fragment_text,
                    _current_matrix,
                    text_matrix,
                    _font_dictionary,
                    font_size,
                ) -> None:
                    if len(text_matrix) >= 6:
                        fragments.append(
                            (
                                text_matrix[4],
                                text_matrix[5],
                                fragment_text,
                                float(font_size or _PDF_DEFAULT_FONT_SIZE),
                            )
                        )

                try:
                    raw_text = page.extract_text(visitor_text=collect_pdf_fragment)
                except TypeError:
                    raw_text = page.extract_text()

                layout_rows = _pdf_layout_rows(fragments)
                text = "\n".join(layout_rows) if layout_rows else raw_text
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
        except Exception as exc:
            logger.warning("PDF extraction failed for %s", filename, exc_info=True)
            raise DocumentExtractionError("Could not extract text from PDF file.") from exc

    elif ext == ".docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)
        except Exception as exc:
            logger.warning("DOCX extraction failed for %s", filename, exc_info=True)
            raise DocumentExtractionError("Could not extract text from DOCX file.") from exc

    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheets_text = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append(" | ".join(cells))
                if rows:
                    sheets_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            wb.close()
            return "\n\n".join(sheets_text)
        except Exception as exc:
            logger.warning("XLSX extraction failed for %s", filename, exc_info=True)
            raise DocumentExtractionError("Could not extract text from XLSX file.") from exc

    else:
        raise ValueError(f"Unsupported file format: {ext}. Supported: PDF, DOCX, TXT, CSV, XLSX")


_VALUE_CONTINUATION_CHARACTERS = {"_", "-", "'", "\u2019"}
_ROLE_TERMINOLOGY = {
    "administrator",
    "analyst",
    "analis",
    "anggota",
    "architect",
    "arsitek",
    "assistant",
    "bendahara",
    "ceo",
    "cfo",
    "chief",
    "ciso",
    "cmo",
    "consultant",
    "coo",
    "coordinator",
    "cto",
    "developer",
    "director",
    "direktur",
    "engineer",
    "executive",
    "founder",
    "head",
    "insinyur",
    "ketua",
    "konsultan",
    "koordinator",
    "lead",
    "leader",
    "manager",
    "manajer",
    "master",
    "member",
    "officer",
    "owner",
    "pemilik",
    "pendiri",
    "pengembang",
    "peneliti",
    "president",
    "researcher",
    "scientist",
    "sekretaris",
    "specialist",
    "spesialis",
    "staff",
    "staf",
    "supervisor",
    "technician",
    "teknisi",
    "vp",
    "data",
    "design",
    "engineering",
    "finance",
    "hr",
    "marketing",
    "operations",
    "product",
    "project",
    "sales",
    "security",
    "software",
    "technical",
    "technology",
    "designer",
    "qa",
    "recruiter",
    "senior",
    "team",
    "ux",
}
_ORGANIZATION_NAME_TERMINOLOGY = {
    "company",
    "department",
    "departemen",
    "division",
    "divisi",
    "group",
    "grup",
    "inc",
    "incorporated",
    "institute",
    "institut",
    "kelompok",
    "labs",
    "lembaga",
    "limited",
    "llc",
    "ltd",
    "organisasi",
    "organisation",
    "organization",
    "perusahaan",
    "perseroan",
    "team",
    "tim",
}
_NAME_CONTEXT_STARTERS = {
    "acts",
    "adalah",
    "as",
    "at",
    "became",
    "bekerja",
    "berperan",
    "bertanggung",
    "bertugas",
    "currently",
    "has",
    "holds",
    "is",
    "joined",
    "menjabat",
    "merupakan",
    "sebagai",
    "serves",
    "was",
    "who",
    "will",
    "works",
    "yang",
}
_ROLE_CONTEXT_STARTERS = {
    "at",
    "di",
    "for",
    "in",
    "of",
    "on",
    "pada",
    "responsible",
    "untuk",
    "with",
    "within",
    "yang",
}
_NAME_ROLE_LINK_TOKENS = _NAME_CONTEXT_STARTERS | {"the"}
_INTRINSICALLY_COMPLETE_ROLES = {"ceo", "cfo", "ciso", "cmo", "coo", "cto"}

_MAX_EXPLICIT_DOCUMENTED_ROSTER = 8
_EXPLICIT_ROSTER_NAME_HEADERS = {"name", "nama"}
_EXPLICIT_ROSTER_ROLE_HEADERS = {"jabatan", "peran", "position", "role"}
_NON_ROSTER_TABLE_FIRST_HEADERS = {
    "action",
    "aksi",
    "decision",
    "hazard",
    "issue",
    "isu",
    "item",
    "keputusan",
    "metric",
    "metrik",
    "requirement",
    "risk",
    "risiko",
    "scenario",
    "skenario",
    "task",
    "tugas",
}
_NON_ROSTER_TABLE_SECOND_HEADERS = {
    "assignee",
    "deadline",
    "description",
    "deskripsi",
    "dampak",
    "due date",
    "impact",
    "likelihood",
    "mitigasi",
    "mitigation",
    "owner",
    "pemilik",
    "pic",
    "priority",
    "status",
}


def _is_non_roster_table_header(fields: list[str]) -> bool:
    return (
        len(fields) >= 2
        and fields[0].casefold() in _NON_ROSTER_TABLE_FIRST_HEADERS
        and fields[1].casefold() in _NON_ROSTER_TABLE_SECOND_HEADERS
    )


def _is_section_heading(value: str) -> bool:
    stripped = value.strip()
    if (
        "|" in stripped
        or len(stripped) > 120
        or stripped.endswith((".", "!", "?", ",", ":", ";"))
    ):
        return False

    letters = [character for character in stripped if character.isalpha()]
    if len(letters) < 5:
        return False
    if all(character.isupper() for character in letters):
        return True

    words = stripped.split()
    if not 1 <= len(words) <= 6:
        return False

    connectors = {"and", "dan", "of", "the", "untuk"}
    for word in words:
        word_letters = [character for character in word if character.isalpha()]
        if word.casefold() in connectors:
            continue
        if (
            not word_letters
            or not word_letters[0].isupper()
            or not all(character.islower() for character in word_letters[1:])
        ):
            return False
    return True


def _semantic_tokens(value: str) -> list[str]:
    tokens: list[str] = []
    current: list[str] = []

    for character in value:
        if character.isalnum():
            current.append(character)
        elif current:
            tokens.append("".join(current).casefold())
            current = []

    if current:
        tokens.append("".join(current).casefold())

    return tokens


def _is_strong_section_heading(value: str) -> bool:
    letters = [character for character in value if character.isalpha()]
    if letters and all(character.isupper() for character in letters):
        return True

    last_word = value.split()[-1] if value.split() else ""
    last_word_letters = [
        character for character in last_word if character.isalpha()
    ]
    return (
        len(last_word_letters) >= 8
        and last_word_letters[-1].casefold() == "s"
    )


def _extract_explicit_documented_roster(
    document_text: str,
) -> list[DocumentedRosterRow]:
    roster: list[DocumentedRosterRow] = []
    seen_names: set[str] = set()
    inside_roster_table = False
    roster_column_count = 0
    pending_section_boundary = False
    pending_section_boundary_is_strong = False

    for raw_line in document_text.splitlines():
        evidence = _normalize_evidence_whitespace(raw_line)
        if not evidence:
            continue

        fields = [
            _normalize_evidence_whitespace(field)
            for field in evidence.split("|")
        ]
        if len(fields) >= 2 and (
            fields[0].casefold() in _EXPLICIT_ROSTER_NAME_HEADERS
            and fields[1].casefold() in _EXPLICIT_ROSTER_ROLE_HEADERS
        ):
            inside_roster_table = True
            roster_column_count = len(fields)
            pending_section_boundary = False
            pending_section_boundary_is_strong = False
            continue

        if inside_roster_table and _is_section_heading(evidence):
            pending_section_boundary = True
            pending_section_boundary_is_strong = (
                _is_strong_section_heading(evidence)
                or roster_column_count == 2
            )
            continue

        if inside_roster_table and _is_non_roster_table_header(fields):
            inside_roster_table = False
            pending_section_boundary = False
            pending_section_boundary_is_strong = False
            continue

        if not inside_roster_table or len(fields) < 2:
            continue

        if pending_section_boundary:
            table_structure_changed = len(fields) != roster_column_count
            if (
                pending_section_boundary_is_strong
                or table_structure_changed
            ):
                inside_roster_table = False
                pending_section_boundary = False
                pending_section_boundary_is_strong = False
                continue
            pending_section_boundary = False
            pending_section_boundary_is_strong = False

        name, role = fields[:2]
        if (
            not name
            or not role
            or len(name) > 100
            or len(role) > 160
            or len(evidence) > 1000
        ):
            continue

        name_tokens = set(_semantic_tokens(name))
        role_tokens = set(_semantic_tokens(role))
        if (
            not name_tokens
            or name_tokens.intersection(_ROLE_TERMINOLOGY)
            or name_tokens.intersection(_ORGANIZATION_NAME_TERMINOLOGY)
            or not role_tokens.intersection(_ROLE_TERMINOLOGY)
        ):
            continue

        normalized_name = name.casefold()
        if normalized_name in seen_names:
            continue

        context = fields[2] if len(fields) >= 3 else ""
        roster.append(
            DocumentedRosterRow(
                name=name,
                role=role,
                context=context,
                evidence=evidence,
            )
        )
        seen_names.add(normalized_name)
        if len(roster) >= _MAX_EXPLICIT_DOCUMENTED_ROSTER:
            break

    return roster


def _is_value_continuation(character: str) -> bool:
    return (
        character.isalnum()
        or character in _VALUE_CONTINUATION_CHARACTERS
    )


def _bounded_occurrences(evidence: str, value: str) -> list[tuple[int, int]]:
    occurrences: list[tuple[int, int]] = []
    search_from = 0

    while True:
        start = evidence.find(value, search_from)
        if start < 0:
            break
        end = start + len(value)
        left_is_clear = start == 0 or not _is_value_continuation(
            evidence[start - 1]
        )
        right_is_clear = end == len(evidence) or not _is_value_continuation(
            evidence[end]
        )
        if left_is_clear and right_is_clear:
            occurrences.append((start, end))
        search_from = start + 1

    return occurrences


def _starts_with_bounded_exact_value(remainder: str, value: str) -> bool:
    if not remainder.startswith(value):
        return False
    end = len(value)
    return end == len(remainder) or not _is_value_continuation(
        remainder[end]
    )


def _has_unverified_trailing_segment(
    evidence: str,
    value_span: tuple[int, int],
    counterpart: str,
    context_starters: set[str],
) -> bool:
    position = value_span[1]
    if position >= len(evidence) or not evidence[position].isspace():
        return False

    while position < len(evidence) and evidence[position].isspace():
        position += 1
    if position >= len(evidence) or not evidence[position].isalnum():
        return False

    remainder = evidence[position:]
    if _starts_with_bounded_exact_value(remainder, counterpart):
        return False

    token_end = 0
    while token_end < len(remainder) and remainder[token_end].isalnum():
        token_end += 1
    next_token = remainder[:token_end].casefold()
    return next_token not in context_starters


def _normalize_evidence_whitespace(value: str) -> str:
    return " ".join(value.split())


def _normalized_evidence_occurrence_count(
    document_text: str,
    evidence: str,
) -> int:
    normalized_document = _normalize_evidence_whitespace(document_text)
    normalized_evidence = _normalize_evidence_whitespace(evidence)
    if not normalized_evidence:
        return 0

    occurrences = 0
    search_from = 0
    while True:
        position = normalized_document.find(normalized_evidence, search_from)
        if position < 0:
            return occurrences
        occurrences += 1
        search_from = position + 1


_MAX_DOCUMENTED_NAME_ROLE_GAP = 240


def _pipe_field_bounds(
    record: str,
    value_span: tuple[int, int],
) -> tuple[int, int] | None:
    if "|" not in record:
        return None

    left_separator = record.rfind("|", 0, value_span[0])
    right_separator = record.find("|", value_span[1])
    field_start = 0 if left_separator < 0 else left_separator + 1
    field_end = len(record) if right_separator < 0 else right_separator
    return field_start, field_end


def _spans_form_complete_name_role_pair(
    record: str,
    name_span: tuple[int, int],
    role_span: tuple[int, int],
) -> bool:
    if max(name_span[0], role_span[0]) < min(name_span[1], role_span[1]):
        return False

    name_field = _pipe_field_bounds(record, name_span)
    role_field = _pipe_field_bounds(record, role_span)
    if name_field is not None and role_field is not None:
        if name_field == role_field:
            return False
        return (
            _normalize_evidence_whitespace(
                record[name_field[0]:name_field[1]]
            )
            == record[name_span[0]:name_span[1]]
            and _normalize_evidence_whitespace(
                record[role_field[0]:role_field[1]]
            )
            == record[role_span[0]:role_span[1]]
        )

    role_value = record[role_span[0]:role_span[1]].casefold()
    if role_value not in _INTRINSICALLY_COMPLETE_ROLES:
        return False

    if name_span[1] <= role_span[0]:
        between = record[name_span[1]:role_span[0]]
    elif role_span[1] <= name_span[0]:
        between = record[role_span[1]:name_span[0]]
    else:
        return False

    return all(
        token in _NAME_ROLE_LINK_TOKENS
        for token in _semantic_tokens(between)
    )


def _has_unique_same_record_document_grounding(
    document_text: str,
    name: str,
    role: str,
    other_values: set[str],
) -> bool:
    normalized_name = _normalize_evidence_whitespace(name)
    normalized_role = _normalize_evidence_whitespace(role)
    normalized_other_values = {
        _normalize_evidence_whitespace(value)
        for value in other_values
        if _normalize_evidence_whitespace(value)
    }
    if not normalized_name or not normalized_role:
        return False

    grounded_records = 0
    for record in document_text.splitlines():
        normalized_record = _normalize_evidence_whitespace(record)
        if not normalized_record:
            continue

        name_spans = _bounded_occurrences(normalized_record, normalized_name)
        role_spans = _bounded_occurrences(normalized_record, normalized_role)
        if len(name_spans) != 1 or not role_spans:
            continue
        if any(
            _bounded_occurrences(normalized_record, value)
            for value in normalized_other_values
        ):
            continue

        name_span = name_spans[0]
        if not any(
            (
                max(
                    name_span[0] - role_span[1],
                    role_span[0] - name_span[1],
                    0,
                )
                <= _MAX_DOCUMENTED_NAME_ROLE_GAP
                and _spans_form_complete_name_role_pair(
                    normalized_record, name_span, role_span
                )
            )
            for role_span in role_spans
        ):
            continue

        grounded_records += 1
        if grounded_records > 1:
            return False

    return grounded_records == 1


def _validate_documented_agent_evidence(
    analysis: DocumentAnalysisResult,
    document_text: str,
) -> None:
    if analysis.team_source is not DocumentTeamSource.DOCUMENTED:
        return

    seen_names: set[str] = set()
    seen_evidence: set[str] = set()

    for agent_index, agent in enumerate(analysis.suggested_agents):
        other_values = {
            value
            for other_index, other_agent in enumerate(analysis.suggested_agents)
            if other_index != agent_index
            for value in (
                other_agent.name, other_agent.role
            )
            if value not in {agent.name, agent.role}
        }
        evidence = agent.source_evidence
        if not evidence:
            raise DocumentRosterEvidenceError("documented team member is missing source_evidence")
        evidence_occurrence_count = _normalized_evidence_occurrence_count(
            document_text, evidence
        )
        same_record_grounded = False
        if evidence_occurrence_count != 1:
            same_record_grounded = _has_unique_same_record_document_grounding(
                document_text,
                agent.name,
                agent.role,
                other_values,
            )
            if not same_record_grounded:
                raise DocumentRosterEvidenceError(
                    "source_evidence must occur exactly once in the document"
                )

        name_spans = _bounded_occurrences(evidence, agent.name)
        role_spans = _bounded_occurrences(evidence, agent.role)
        if len(name_spans) != 1 or len(role_spans) != 1:
            raise DocumentRosterEvidenceError(
                "source_evidence must contain one bounded exact name and role"
            )

        name_start, name_end = name_spans[0]
        role_start, role_end = role_spans[0]
        if max(name_start, role_start) < min(name_end, role_end):
            raise DocumentRosterEvidenceError(
                "name and role overlap in source_evidence"
            )

        name_tokens = set(_semantic_tokens(agent.name))
        if name_tokens.intersection(_ROLE_TERMINOLOGY):
            raise DocumentRosterEvidenceError(
                "documented name contains role terminology"
            )
        if name_tokens.intersection(_ORGANIZATION_NAME_TERMINOLOGY):
            raise DocumentRosterEvidenceError(
                "documented name contains organization terminology"
            )

        if _has_unverified_trailing_segment(
            evidence,
            name_spans[0],
            agent.role,
            _NAME_CONTEXT_STARTERS,
        ):
            raise DocumentRosterEvidenceError(
                "documented name appears incomplete in source_evidence"
            )
        if (
            not same_record_grounded
            and _has_unverified_trailing_segment(
                evidence,
                role_spans[0],
                agent.name,
                _ROLE_CONTEXT_STARTERS,
            )
        ):
            raise DocumentRosterEvidenceError(
                "documented role appears incomplete in source_evidence"
            )

        normalized_name = agent.name.casefold()
        if normalized_name in seen_names or evidence in seen_evidence:
            raise DocumentRosterEvidenceError(
                "documented team evidence is duplicated"
            )
        seen_names.add(normalized_name)
        seen_evidence.add(evidence)


def _validate_explicit_roster_completeness(
    analysis: DocumentAnalysisResult,
    explicit_roster: list[DocumentedRosterRow],
) -> None:
    if not explicit_roster:
        return

    if analysis.team_source is not DocumentTeamSource.DOCUMENTED:
        raise DocumentRosterEvidenceError(
            "explicit table roster requires team_source=documented"
        )

    expected_pairs = {
        (row.name, row.role)
        for row in explicit_roster
    }
    returned_pairs = {
        (agent.name, agent.role)
        for agent in analysis.suggested_agents
    }
    missing_pairs = expected_pairs - returned_pairs
    unexpected_pairs = returned_pairs - expected_pairs
    if not missing_pairs and not unexpected_pairs:
        return

    mismatch_details = []
    if missing_pairs:
        mismatch_details.append(
            f"omitted {len(missing_pairs)} explicit team members"
        )
    if unexpected_pairs:
        mismatch_details.append(
            f"included {len(unexpected_pairs)} non-table team members"
        )
    raise DocumentRosterEvidenceError(
        "documented table roster " + " and ".join(mismatch_details)
    )


def _complete_explicit_documented_roster(
    analysis: DocumentAnalysisResult,
    explicit_roster: list[DocumentedRosterRow],
) -> DocumentAnalysisResult:
    existing_agents = {
        (agent.name, agent.role): agent
        for agent in analysis.suggested_agents
    }
    completed_agents = []

    for row in explicit_roster:
        existing_agent = existing_agents.get((row.name, row.role))
        if existing_agent is not None:
            agent_payload = existing_agent.model_dump()
            agent_payload["source_evidence"] = row.evidence
        else:
            agent_payload = {
                "name": row.name,
                "role": row.role,
                "source_evidence": row.evidence,
                "type": (
                    row.context[:160]
                    or "Documented team member"
                ),
                "rationale": (
                    "Explicitly listed in the uploaded document as "
                    f"{row.role}."
                ),
                "personality": {
                    "empathy": 50,
                    "ambition": 50,
                    "stressTolerance": 50,
                    "agreeableness": 50,
                    "assertiveness": 50,
                },
            }
        completed_agents.append(agent_payload)

    payload = analysis.model_dump()
    payload["team_source"] = DocumentTeamSource.DOCUMENTED
    payload["suggested_agents"] = completed_agents
    return DocumentAnalysisResult.model_validate(payload)


_ROSTER_NOT_APPLIED_INSIGHT = (
    "Team roster was not auto-applied because its document evidence "
    "could not be verified."
)


def _without_verified_roster(
    analysis: DocumentAnalysisResult,
) -> DocumentAnalysisResult:
    payload = analysis.model_dump()
    payload["team_source"] = DocumentTeamSource.NONE
    payload["suggested_agents"] = []

    insights = list(payload["actionable_insights"])
    if _ROSTER_NOT_APPLIED_INSIGHT not in insights:
        if len(insights) >= 12:
            insights[-1] = _ROSTER_NOT_APPLIED_INSIGHT
        else:
            insights.append(_ROSTER_NOT_APPLIED_INSIGHT)
    payload["actionable_insights"] = insights

    return DocumentAnalysisResult.model_validate(payload)


def _with_corrected_roster(
    initial: DocumentAnalysisResult,
    corrected: DocumentAnalysisResult,
) -> DocumentAnalysisResult:
    if corrected.team_source is DocumentTeamSource.NONE:
        return _without_verified_roster(initial)

    payload = initial.model_dump()
    payload["team_source"] = corrected.team_source
    payload["suggested_agents"] = [
        agent.model_dump()
        for agent in corrected.suggested_agents
    ]
    return DocumentAnalysisResult.model_validate(payload)


def _public_document_analysis(analysis: DocumentAnalysisResult) -> dict:
    return analysis.model_dump(
        mode="json",
        exclude={
            "suggested_agents": {"__all__": {"source_evidence"}},
        },
    )


async def analyze_document(text: str, filename: str) -> dict:
    """
    Use LLM to analyze extracted document text and generate structured output
    for simulation setup assistance.
    """
    from services.llm_service import _dispatch_llm_call, LLM_PROVIDER

    # Truncate text to avoid token limits (roughly 15k chars)
    truncated = text[:15000]
    if len(text) > 15000:
        truncated += "\n\n[... document truncated for analysis ...]"
    explicit_roster = _extract_explicit_documented_roster(truncated)
    explicit_roster_listing = "\n".join(
        f"- {row.name} | {row.role}"
        for row in explicit_roster
    )
    explicit_roster_prompt = ""
    if explicit_roster_listing:
        explicit_roster_prompt = f"""\nEXPLICIT NAME/ROLE TABLE ROSTER:
The parser found {len(explicit_roster)} documented members below. Return every one of them exactly once in suggested_agents; do not select only a subset.
{explicit_roster_listing}
"""

    system_prompt = """You are an expert organizational analyst AI. You analyze business documents to extract useful information for team dynamics simulation.

Given a document, extract and structure the following information. Respond ONLY with valid JSON:
{
  "company_name": "The company or organization name mentioned in the document. If not explicitly stated, infer a suitable name from context.",
  "company_culture": "1-2 sentence description of the company culture, work environment, or organizational context extracted from the document.",
  "operating_context": "A specific description of how the organization operates, including its business context, constraints, and current pressures.",
  "summary": "A 2-3 sentence summary of what this document is about",
  "key_requirements": ["requirement 1", "requirement 2", "requirement 3"],
  "team_risks": ["risk 1", "risk 2", "risk 3"],
  "suggested_crisis": {
    "title": "A crisis scenario title derived from the document",
    "description": "1-2 sentence description of a realistic crisis that could emerge from this context"
  },
  "team_source": "documented, inferred, or none based on the evidence for team members in the document",
  "suggested_agents": [
    {
      "name": "For documented teams, the complete name copied verbatim from the document; for inferred teams, a concise generated name",
      "role": "Suggested team role (e.g., Tech Lead, Product Manager)",
      "source_evidence": "For documented teams, a unique short verbatim sentence or list entry containing the exact complete name and exact complete role in any order; for inferred teams, null",
      "type": "Personality type descriptor (e.g., Ambitious & Driven, Strict & Methodical)",
      "rationale": "Why this role is needed based on the document",
      "personality": {
        "empathy": 50,
        "ambition": 50,
        "stressTolerance": 50,
        "agreeableness": 50,
        "assertiveness": 50
      }
    }
  ],
  "suggested_team_rules": ["rule 1", "rule 2", "rule 3"],
  "actionable_insights": ["insight 1", "insight 2", "insight 3"]
}

Use team_source="documented" only when the document names people or explicitly lists team members.
Use team_source="inferred" when the document clearly implies necessary roles but does not name people.
Use team_source="none" and suggested_agents=[] when team evidence is insufficient.
Never present inferred names as names found in the document.
When team_source="documented", copy every person's complete name exactly as written, including all surname/family-name tokens.
Never shorten a documented full name (for example, return "Avery Chen", not "Avery").
Always produce a specific company name, culture, operating context, and realistic pressure scenario grounded in the document.
For every documented person, source_evidence must be copied verbatim and occur only once in the document.
Use a short complete sentence or list entry containing the exact complete name and exact complete job role as separate text. Either order and surrounding context are allowed; never shorten either value.
For inferred team members, always return source_evidence=null.

Rules:
1. Extract SPECIFIC information from the document, not generic advice
2. Team risks should relate to actual content in the document
3. The suggested crisis should be realistic and derived from document context
4. For team_source="documented", return every explicitly listed person, up to 8 team members; For team_source="inferred", suggest 2-4 relevant roles; when team_source is "none", return suggested_agents=[]
5. Keep all text concise and actionable
6. If the document is not related to business/team/project, still try to extract useful team simulation insights
7. For company_name: extract the actual company name from the document. If no company name is found, create a meaningful name based on the document's industry/context
8. For company_culture: describe the work environment and organizational dynamics based on what the document reveals
9. For each suggested agent, assign personality trait values (0-100) that realistically match their role and type. A Tech Lead might have high assertiveness, a Junior Dev might have high ambition but low stress tolerance, etc.
10. For suggested_team_rules: extract or infer 2-4 team operating rules/norms that would be relevant based on the document context"""

    user_prompt = f"""Analyze this document for team dynamics simulation setup:

FILENAME: {filename}

{explicit_roster_prompt}
DOCUMENT CONTENT:
{truncated}

Generate structured analysis. Pay special attention to extracting the company/organization name, describing the company culture, and suggesting realistic team members with appropriate personality traits."""

    try:
        result = await _dispatch_llm_call(
            system_prompt,
            user_prompt,
            LLM_PROVIDER,
            max_tokens=2000,
            allow_model_fallback=False,
            response_model=DocumentAnalysisResult,
        )
        analysis = DocumentAnalysisResult.model_validate(result)

        try:
            _validate_documented_agent_evidence(analysis, truncated)
            _validate_explicit_roster_completeness(analysis, explicit_roster)
        except DocumentRosterEvidenceError as evidence_error:
            logger.warning(
                "Document roster evidence needs one corrective retry for %s: %s",
                filename,
                evidence_error,
            )
            corrective_user_prompt = f"""CORRECTIVE RETRY: The previous analysis was schema-valid, but its documented team roster evidence could not be verified.

Preserve company_name, company_culture, operating_context, summary, suggested_crisis, and all non-roster insights. Correct only team_source and suggested_agents as needed.

For each documented agent:
- copy the complete name and complete role exactly from the document
- provide a short source_evidence sentence or list entry containing both
- source_evidence must match one unique document passage; whitespace runs may differ because of PDF extraction
- never invent, shorten, or reconstruct a person's name

If no uniquely grounded roster can be returned, use team_source="none" and suggested_agents=[].

{explicit_roster_prompt}
PREVIOUS SCHEMA-VALID ANALYSIS:
{analysis.model_dump_json()}

DOCUMENT CONTENT:
{truncated}

Return the complete corrected structured analysis."""

            corrected_result = await _dispatch_llm_call(
                system_prompt,
                corrective_user_prompt,
                LLM_PROVIDER,
                max_tokens=2000,
                allow_model_fallback=False,
                response_model=DocumentAnalysisResult,
            )
            corrected_analysis = DocumentAnalysisResult.model_validate(
                corrected_result
            )
            try:
                _validate_documented_agent_evidence(
                    corrected_analysis,
                    truncated,
                )
                _validate_explicit_roster_completeness(
                    corrected_analysis, explicit_roster
                )
            except DocumentRosterEvidenceError as corrected_evidence_error:
                logger.warning(
                    "Document roster evidence remained unverifiable for %s: %s",
                    filename,
                    corrected_evidence_error,
                )
                if explicit_roster:
                    completed_roster = _complete_explicit_documented_roster(
                        corrected_analysis,
                        explicit_roster,
                    )
                    _validate_documented_agent_evidence(
                        completed_roster,
                        truncated,
                    )
                    _validate_explicit_roster_completeness(
                        completed_roster,
                        explicit_roster,
                    )
                    merged_analysis = _with_corrected_roster(
                        analysis,
                        completed_roster,
                    )
                    return _public_document_analysis(merged_analysis)

                return _public_document_analysis(
                    _without_verified_roster(analysis)
                )
            merged_analysis = _with_corrected_roster(
                analysis, corrected_analysis
            )
            return _public_document_analysis(merged_analysis)

        return _public_document_analysis(analysis)
    except Exception as exc:
        logger.exception("Document analysis failed for %s", filename)
        raise DocumentAnalysisError(
            "AI could not analyze this document. Please retry with a clearer text-based file."
        ) from exc
